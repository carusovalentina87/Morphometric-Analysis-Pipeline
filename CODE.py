#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Morphometric Analysis Pipeline for Archaeological Bitumen-based Samples
====================================

This script:
- Reads binary (or near-binary) TIFF images from an input folder
- Extracts pore/feature contours (external)
- Computes morphometric properties (area, major/minor axes, aspect ratio)
- Exports:
  1) One Excel workbook with one sheet per image
  2) One figure per image (image + histograms)

Outputs are saved inside a newly created `output/` folder within the chosen output directory.

Notes
-----
- Images are expected to be binary (0/255). If not, they are binarized with a fixed threshold at 128.
- Spatial calibration uses DPI: mm_per_pixel = 25.4 / DPI
"""

from __future__ import annotations

import os
import glob
import argparse
from pathlib import Path

import cv2
import numpy as np
import matplotlib.pyplot as plt
from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import Font


# ============================================================
# Default parameters
# ============================================================

DEFAULT_DPI = 72
DEFAULT_DDOF_STD = 1
DEFAULT_HIST_BINS = 30


# ============================================================
# Image processing functions
# ============================================================

def verify_and_binarize(img_path: str) -> np.ndarray:
    """
    Verify whether an image is binary (0 / 255). If not, apply thresholding at 128.

    Parameters
    ----------
    img_path : str
        Path to the input image.

    Returns
    -------
    numpy.ndarray
        Binary image (uint8).
    """
    img = Image.open(img_path).convert("L")
    img_np = np.array(img)

    unique_vals = np.unique(img_np)
    if set(unique_vals).issubset({0, 255}):
        return img_np.astype(np.uint8)

    _, img_bin = cv2.threshold(img_np, 128, 255, cv2.THRESH_BINARY)
    return img_bin.astype(np.uint8)


def compute_properties(img_bin: np.ndarray, mm_per_pixel: float):
    """
    Compute morphological properties for each detected feature.

    Parameters
    ----------
    img_bin : numpy.ndarray
        Binary image.
    mm_per_pixel : float
        Conversion factor from pixel to mm.

    Returns
    -------
    list of dict
        Morphological properties for each feature.
    list
        Detected contours.
    """
    results = []
    contours, _ = cv2.findContours(
        img_bin, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE
    )

    for cnt in contours:
        area_px = cv2.contourArea(cnt)
        if area_px <= 0:
            continue

        area_mm2 = area_px * (mm_per_pixel ** 2)
        rect = cv2.minAreaRect(cnt)
        w_px, h_px = rect[1]

        major = max(w_px, h_px) * mm_per_pixel
        minor = min(w_px, h_px) * mm_per_pixel
        aspect = major / minor if minor > 0 else np.nan

        results.append(
            {
                "Area (mm²)": area_mm2,
                "Major Axis Length (mm)": major,
                "Minor Axis Length (mm)": minor,
                "Aspect Ratio": aspect,
            }
        )

    return results, contours


def plot_image_and_histograms(
    img_bin: np.ndarray,
    contours,
    results,
    output_path: str,
    hist_bins: int
) -> None:
    """
    Generate a combined figure with the binary image and histograms
    of morphological properties.

    Parameters
    ----------
    img_bin : numpy.ndarray
        Binary image.
    contours : list
        Feature contours.
    results : list of dict
        Extracted morphological properties.
    output_path : str
        Path for the output TIFF figure.
    hist_bins : int
        Number of histogram bins.
    """
    areas = [r["Area (mm²)"] for r in results]
    major_axes = [r["Major Axis Length (mm)"] for r in results]
    minor_axes = [r["Minor Axis Length (mm)"] for r in results]
    aspect_ratios = [r["Aspect Ratio"] for r in results]

    img_rgb = cv2.cvtColor(img_bin, cv2.COLOR_GRAY2BGR)
    cv2.drawContours(img_rgb, contours, -1, (0, 0, 0), 2)

    fig = plt.figure(figsize=(16, 8))

    ax_img = plt.subplot2grid((2, 3), (0, 0), rowspan=2)
    ax_img.imshow(cv2.cvtColor(img_rgb, cv2.COLOR_BGR2RGB))
    ax_img.axis("off")

    ax1 = plt.subplot2grid((2, 3), (0, 1))
    ax1.hist(areas, bins=hist_bins)
    ax1.set_xlabel("Area (mm²)", fontsize=14)
    ax1.set_ylabel("Frequency", fontsize=14)

    ax2 = plt.subplot2grid((2, 3), (0, 2))
    ax2.hist(major_axes, bins=hist_bins)
    ax2.set_xlabel("Major Axis Length (mm)", fontsize=14)
    ax2.set_ylabel("Frequency", fontsize=14)

    ax3 = plt.subplot2grid((2, 3), (1, 1))
    ax3.hist(minor_axes, bins=hist_bins)
    ax3.set_xlabel("Minor Axis Length (mm)", fontsize=14)
    ax3.set_ylabel("Frequency", fontsize=14)

    ax4 = plt.subplot2grid((2, 3), (1, 2))
    ax4.hist(aspect_ratios, bins=hist_bins)
    ax4.set_xlabel("Aspect Ratio", fontsize=14)
    ax4.set_ylabel("Frequency", fontsize=14)

    plt.tight_layout()
    fig.savefig(output_path, dpi=300, format="tiff")
    plt.close(fig)


# ============================================================
# Main execution
# ============================================================

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Morphometric analysis of binary TIFF images with Excel + figure export."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="Input folder containing .tif/.tiff images."
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Output base folder. If omitted, uses the input folder."
    )
    parser.add_argument(
        "--dpi",
        type=float,
        default=DEFAULT_DPI,
        help=f"Image DPI used for mm conversion (default: {DEFAULT_DPI})."
    )
    parser.add_argument(
        "--ddof",
        type=int,
        choices=[0, 1],
        default=DEFAULT_DDOF_STD,
        help="Std ddof: 1=sample std, 0=population std (default: 1)."
    )
    parser.add_argument(
        "--hist-bins",
        type=int,
        default=DEFAULT_HIST_BINS,
        help=f"Number of bins for histograms (default: {DEFAULT_HIST_BINS})."
    )

    args = parser.parse_args()

    input_dir = Path(args.input)
    if not input_dir.exists():
        raise FileNotFoundError(f"Input folder not found: {input_dir}")

    base_output_dir = Path(args.output) if args.output else input_dir
    output_dir = base_output_dir / "output"
    output_dir.mkdir(parents=True, exist_ok=True)

    mm_per_pixel = 25.4 / float(args.dpi)

    workbook = Workbook()
    bold_font = Font(name="Times New Roman", size=12, bold=True)
    normal_font = Font(name="Times New Roman", size=12)

    image_files = (
        glob.glob(str(input_dir / "*.tif")) +
        glob.glob(str(input_dir / "*.tiff"))
    )

    if not image_files:
        print("No TIFF images found in:", str(input_dir))
        return

    for file_path in image_files:
        file_path = str(file_path)
        filename = os.path.basename(file_path)
        print(f"Processing: {filename}")

        img_bin = verify_and_binarize(file_path)
        results, contours = compute_properties(img_bin, mm_per_pixel)

        # ----------------------------------------------------
        # Excel export (one sheet per image)
        # ----------------------------------------------------
        sheet_name = os.path.splitext(filename)[0][:31]
        sheet = workbook.create_sheet(title=sheet_name)

        sheet.append(["Number of Features"])
        sheet.append([len(results)])

        header = [
            "Area (mm²)", "Area Std",
            "Major Axis Length (mm)", "Major Axis Std",
            "Minor Axis Length (mm)", "Minor Axis Std",
            "Aspect Ratio", "Aspect Ratio Std",
        ]
        sheet.append(header)
        for cell in sheet[3]:
            cell.font = bold_font

        keys = [
            "Area (mm²)",
            "Major Axis Length (mm)",
            "Minor Axis Length (mm)",
            "Aspect Ratio",
        ]

        stds = {}
        for k in keys:
            vals = np.array([r[k] for r in results], dtype=float)
            stds[k] = np.std(vals, ddof=args.ddof) if len(vals) > 1 else np.nan

        for r in results:
            row = [
                r["Area (mm²)"], stds["Area (mm²)"],
                r["Major Axis Length (mm)"], stds["Major Axis Length (mm)"],
                r["Minor Axis Length (mm)"], stds["Minor Axis Length (mm)"],
                r["Aspect Ratio"], stds["Aspect Ratio"],
            ]
            sheet.append(row)
            for cell in sheet[sheet.max_row]:
                cell.font = normal_font

        # ----------------------------------------------------
        # Figure export
        # ----------------------------------------------------
        base_name = os.path.splitext(filename)[0]
        out_fig = str(output_dir / f"{base_name}_graphs.tiff")
        plot_image_and_histograms(
            img_bin=img_bin,
            contours=contours,
            results=results,
            output_path=out_fig,
            hist_bins=args.hist_bins
        )

    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    out_xlsx = output_dir / "morphometric_analysis_result.xlsx"
    workbook.save(str(out_xlsx))

    print("Analysis completed.")
    print("Outputs saved to:", str(output_dir))
    print("Excel saved:", str(out_xlsx))


if __name__ == "__main__":
    main()

